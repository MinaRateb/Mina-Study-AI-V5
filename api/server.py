import os, sqlite3, secrets
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env'))
except ImportError:
    pass
from datetime import datetime
from functools import wraps
from flask import Flask, request, jsonify, session, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from openai import OpenAI

BASE=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB=os.path.join(BASE,"mina_study.db")
app=Flask(__name__,static_folder=os.path.join(BASE,"public"),static_url_path="")
app.secret_key=os.environ.get("SESSION_SECRET",secrets.token_hex(32))
app.config.update(SESSION_COOKIE_HTTPONLY=True,SESSION_COOKIE_SAMESITE="Lax",SESSION_COOKIE_SECURE=os.environ.get('COOKIE_SECURE','0')=='1')
CSRF_TOKEN = secrets.token_urlsafe(32)
OPENAI_API_KEY=os.environ.get("OPENAI_API_KEY","").strip()
client=OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None
MODEL=os.environ.get("OPENAI_MODEL","gpt-5.6-luna")

DATABASE_URL=os.environ.get("DATABASE_URL","").strip()

class DBWrap:
    def __init__(self, conn, postgres=False):
        self.conn=conn; self.postgres=postgres
    def execute(self, sql, params=()):
        if self.postgres:
            sql=sql.replace("?", "%s")
        return self.conn.execute(sql, params)
    def commit(self): self.conn.commit()
    def rollback(self): self.conn.rollback()
    def close(self): self.conn.close()
    def executescript(self, script):
        if self.postgres:
            with self.conn.cursor() as cur:
                for stmt in script.split(';'):
                    stmt=stmt.strip()
                    if stmt: cur.execute(stmt)
        else:
            self.conn.executescript(script)

def db():
    if DATABASE_URL:
        try:
            import psycopg
            from psycopg.rows import dict_row
            conn=psycopg.connect(DATABASE_URL, row_factory=dict_row)
            return DBWrap(conn, postgres=True)
        except Exception as e:
            raise RuntimeError(f"ØªØ¹Ø°Ø± Ø§Ù„Ø§ØªØµØ§Ù„ Ø¨Ù‚Ø§Ø¹Ø¯Ø© PostgreSQL: {e}")
    c=sqlite3.connect(DB); c.row_factory=sqlite3.Row; return DBWrap(c, postgres=False)

def init_db():
    c=db()
    if c.postgres:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS users(id BIGSERIAL PRIMARY KEY,name TEXT NOT NULL,email TEXT UNIQUE NOT NULL,password_hash TEXT NOT NULL,role TEXT NOT NULL DEFAULT 'student',status TEXT NOT NULL DEFAULT 'pending',created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS profiles(user_id BIGINT PRIMARY KEY,stage TEXT,track TEXT,location TEXT,notes TEXT);
        CREATE TABLE IF NOT EXISTS schedules(id BIGSERIAL PRIMARY KEY,user_id BIGINT NOT NULL,day TEXT,title TEXT,start_time TEXT,end_time TEXT,type TEXT);
        CREATE TABLE IF NOT EXISTS grades(id BIGSERIAL PRIMARY KEY,user_id BIGINT NOT NULL,subject TEXT,score DOUBLE PRECISION,max_score DOUBLE PRECISION,weak_point TEXT,created_at TEXT);
        """)
    else:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT NOT NULL,email TEXT UNIQUE NOT NULL,password_hash TEXT NOT NULL,role TEXT NOT NULL DEFAULT 'student',status TEXT NOT NULL DEFAULT 'pending',created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS profiles(user_id INTEGER PRIMARY KEY,stage TEXT,track TEXT,location TEXT,notes TEXT);
        CREATE TABLE IF NOT EXISTS schedules(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER NOT NULL,day TEXT,title TEXT,start_time TEXT,end_time TEXT,type TEXT);
        CREATE TABLE IF NOT EXISTS grades(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER NOT NULL,subject TEXT,score REAL,max_score REAL,weak_point TEXT,created_at TEXT);
        """)
    c.commit(); c.close()
def login_required(f):
    @wraps(f)
    def w(*a,**k):
        if not session.get("uid"): return jsonify(error="ØªØ³Ø¬ÙŠÙ„ Ø§Ù„Ø¯Ø®ÙˆÙ„ Ù…Ø·Ù„ÙˆØ¨"),401
        return f(*a,**k)
    return w
def admin_required(f):
    @wraps(f)
    def w(*a,**k):
        if session.get("role")!="admin": return jsonify(error="ØµÙ„Ø§Ø­ÙŠØ© Ø§Ù„Ù…Ø³Ø¤ÙˆÙ„ Ù…Ø·Ù„ÙˆØ¨Ø©"),403
        return f(*a,**k)
    return w
@app.get("/api/csrf")
def csrf():
    return jsonify(token=CSRF_TOKEN)

def csrf_required(f):
    @wraps(f)
    def w(*a,**k):
        if request.headers.get("X-CSRF-Token") != CSRF_TOKEN:
            return jsonify(error="CSRF token missing or invalid"),403
        return f(*a,**k)
    return w

@app.get("/")
def home(): return send_from_directory(os.path.join(BASE,"public"),"index.html")
@app.post("/api/signup")
@csrf_required
def signup():
    d=request.get_json() or {}; name=d.get("name","").strip(); email=d.get("email","").strip().lower(); pw=d.get("password","")
    if not name or not email or len(pw)<8: return jsonify(error="Ø§ÙƒØªØ¨ Ø§Ù„Ø§Ø³Ù… ÙˆØ§Ù„Ø¥ÙŠÙ…ÙŠÙ„ ÙˆÙƒÙ„Ù…Ø© Ù…Ø±ÙˆØ± 8 Ø£Ø­Ø±Ù Ø¹Ù„Ù‰ Ø§Ù„Ø£Ù‚Ù„"),400
    c=db()
    try:
        c.execute("INSERT INTO users(name,email,password_hash,created_at) VALUES(?,?,?,?)",(name,email,generate_password_hash(pw),datetime.utcnow().isoformat()))
        c.commit()
    except sqlite3.IntegrityError: c.close(); return jsonify(error="Ø§Ù„Ø¥ÙŠÙ…ÙŠÙ„ Ù…Ø³ØªØ®Ø¯Ù… Ø¨Ø§Ù„ÙØ¹Ù„"),409
    c.close(); return jsonify(message="ØªÙ… Ø§Ù„ØªØ³Ø¬ÙŠÙ„. Ø§Ù†ØªØ¸Ø± Ù…ÙˆØ§ÙÙ‚Ø© Ø§Ù„Ù…Ø³Ø¤ÙˆÙ„.")
@app.post("/api/login")
@csrf_required
def login():
    d=request.get_json() or {}; c=db(); u=c.execute("SELECT * FROM users WHERE email=?",(d.get("email","").lower(),)).fetchone(); c.close()
    if not u or not check_password_hash(u["password_hash"],d.get("password","")): return jsonify(error="Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø¯Ø®ÙˆÙ„ ØºÙŠØ± ØµØ­ÙŠØ­Ø©"),401
    if u["status"]!="approved": return jsonify(error="Ø§Ù„Ø­Ø³Ø§Ø¨ ÙÙŠ Ø§Ù†ØªØ¸Ø§Ø± Ù…ÙˆØ§ÙÙ‚Ø© Ø§Ù„Ù…Ø³Ø¤ÙˆÙ„"),403
    session["uid"]=u["id"]; session["role"]=u["role"]; session["name"]=u["name"]; return jsonify(name=u["name"],role=u["role"])
@app.post("/api/logout")
@csrf_required
def logout(): session.clear(); return jsonify(ok=True)
@app.get("/api/me")
@login_required
def me():
    c=db(); u=c.execute("SELECT id,name,email,role,status,created_at FROM users WHERE id=?",(session["uid"],)).fetchone()
    p=c.execute("SELECT * FROM profiles WHERE user_id=?",(session["uid"],)).fetchone()
    s=c.execute("SELECT id,user_id,day,title,start_time AS start,end_time AS end,type FROM schedules WHERE user_id=? ORDER BY id",(session["uid"],)).fetchall()
    g=c.execute("SELECT * FROM grades WHERE user_id=? ORDER BY id DESC",(session["uid"],)).fetchall(); c.close()
    return jsonify(user=dict(u),profile=dict(p) if p else {},schedules=[dict(x) for x in s],grades=[dict(x) for x in g])
@app.post("/api/profile")
@csrf_required
@login_required
def profile():
    d=request.get_json() or {}; c=db(); c.execute("""INSERT INTO profiles(user_id,stage,track,location,notes) VALUES(?,?,?,?,?)
    ON CONFLICT(user_id) DO UPDATE SET stage=excluded.stage,track=excluded.track,location=excluded.location,notes=excluded.notes""",(session["uid"],d.get("stage",""),d.get("track",""),d.get("location",""),d.get("notes",""))); c.commit(); c.close(); return jsonify(ok=True)
@app.post("/api/schedule")
@csrf_required
@login_required
def schedule():
    d=request.get_json() or {}; c=db(); c.execute("INSERT INTO schedules(user_id,day,title,start_time,end_time,type) VALUES(?,?,?,?,?,?)",(session["uid"],d.get("day",""),d.get("title",""),d.get("start",""),d.get("end",""),d.get("type","Ø¯Ø±Ø³"))); c.commit(); c.close(); return jsonify(ok=True)
@app.post("/api/grade")
@csrf_required
@login_required
def grade():
    d=request.get_json() or {}; c=db(); c.execute("INSERT INTO grades(user_id,subject,score,max_score,weak_point,created_at) VALUES(?,?,?,?,?,?)",(session["uid"],d.get("subject",""),d.get("score",0),d.get("max_score",100),d.get("weak_point",""),datetime.utcnow().isoformat())); c.commit(); c.close(); return jsonify(ok=True)

@app.get("/api/admin/users")
@admin_required
def admin_users():
    c=db()
    rows=c.execute("SELECT id,name,email,role,status,created_at FROM users ORDER BY id DESC").fetchall()
    c.close()
    return jsonify(users=[dict(x) for x in rows])

@app.post("/api/admin/set-role/<int:uid>")
@admin_required
@csrf_required
def set_role(uid):
    d=request.get_json() or {}
    role=d.get("role")
    if role not in ("student","admin"):
        return jsonify(error="Ø§Ù„Ø¯ÙˆØ± ØºÙŠØ± ØµØ­ÙŠØ­"),400
    if uid==session.get("uid") and role!="admin":
        return jsonify(error="Ù„Ø§ ÙŠÙ…ÙƒÙ†Ùƒ Ø¥Ø²Ø§Ù„Ø© ØµÙ„Ø§Ø­ÙŠØ© Ø§Ù„Ù…Ø³Ø¤ÙˆÙ„ Ù…Ù† Ù†ÙØ³Ùƒ"),400
    c=db()
    target=c.execute("SELECT id,role FROM users WHERE id=?",(uid,)).fetchone()
    if not target:
        c.close(); return jsonify(error="Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯"),404
    c.execute("UPDATE users SET role=? WHERE id=?",(role,uid))
    c.commit(); c.close()
    return jsonify(ok=True)

@app.post("/api/admin/set-status/<int:uid>")
@admin_required
@csrf_required
def set_status(uid):
    d=request.get_json() or {}
    status=d.get("status")
    if status not in ("approved","pending","blocked"):
        return jsonify(error="Ø§Ù„Ø­Ø§Ù„Ø© ØºÙŠØ± ØµØ­ÙŠØ­Ø©"),400
    if uid==session.get("uid") and status!="approved":
        return jsonify(error="Ù„Ø§ ÙŠÙ…ÙƒÙ†Ùƒ ØªØ¹Ø·ÙŠÙ„ Ø­Ø³Ø§Ø¨Ùƒ Ø¨Ù†ÙØ³Ùƒ"),400
    c=db(); c.execute("UPDATE users SET status=? WHERE id=?",(status,uid)); c.commit(); c.close()
    return jsonify(ok=True)

@app.get("/api/admin/pending")
@admin_required
def pending():
    c=db(); rows=c.execute("SELECT id,name,email,status,created_at FROM users WHERE status='pending' ORDER BY id").fetchall(); c.close(); return jsonify(users=[dict(x) for x in rows])
@app.post("/api/admin/approve/<int:uid>")
@csrf_required
@admin_required
def approve(uid):
    c=db(); c.execute("UPDATE users SET status='approved' WHERE id=?",(uid,)); c.commit(); c.close(); return jsonify(ok=True)
@app.post("/api/admin/reject/<int:uid>")
@csrf_required
@admin_required
def reject(uid):
    c=db(); c.execute("DELETE FROM users WHERE id=?",(uid,)); c.commit(); c.close(); return jsonify(ok=True)

SYSTEM="""You are Mina Study AI, a study assistant for any student. Use only the current student's supplied data.
Help explain lessons step by step, practice, analyze mistakes, plan study sessions, and recommend useful AI tools.
When asked for educational videos, suggest reliable search phrases and types of videos; do not invent URLs.
When recommending tools, say what each tool is good for and prefer official resources when possible.
Use Arabic when the student writes Arabic. Be encouraging and concise. Do not provide medical advice.
"""
@app.post("/api/ai")
@csrf_required
@login_required
def ai():
    if client is None:
        return jsonify(error="Ù„Ù… ÙŠØªÙ… Ø¥Ø¹Ø¯Ø§Ø¯ OPENAI_API_KEY Ø¹Ù„Ù‰ Ø§Ù„Ø³ÙŠØ±ÙØ± Ø¨Ø¹Ø¯. Ø§Ù„Ù…ÙˆÙ‚Ø¹ Ù†ÙØ³Ù‡ ÙŠØ¹Ù…Ù„ØŒ Ù„ÙƒÙ† Ù…Ø³Ø§Ø¹Ø¯ AI ÙŠØ­ØªØ§Ø¬ Ù…ÙØªØ§Ø­ API."),503
    d=request.get_json() or {}; c=db()
    u=c.execute("SELECT name FROM users WHERE id=?",(session["uid"],)).fetchone()
    p=c.execute("SELECT * FROM profiles WHERE user_id=?",(session["uid"],)).fetchone()
    s=c.execute("SELECT id,user_id,day,title,start_time AS start,end_time AS end,type FROM schedules WHERE user_id=?",(session["uid"],)).fetchall()
    g=c.execute("SELECT subject,score,max_score,weak_point FROM grades WHERE user_id=? ORDER BY id DESC LIMIT 20",(session["uid"],)).fetchall(); c.close()
    ctx=f"""Student: {u['name']}
Profile: {dict(p) if p else {}}
Schedule: {[dict(x) for x in s]}
Grades: {[dict(x) for x in g]}
Request type: {d.get('type','')}
Subject: {d.get('subject','')}
Message: {d.get('message','')}"""
    try:
        r=client.responses.create(model=MODEL,input=[{"role":"system","content":SYSTEM},{"role":"user","content":ctx}])
        return jsonify(answer=r.output_text)
    except Exception as e: return jsonify(error=str(e)),500

@app.post("/api/import-plan")
@login_required
@csrf_required
def import_plan():
    """Import a student weekly plan from Arabic/English CSV or XLSX."""
    if "file" not in request.files:
        return jsonify(error="Ø§Ø±ÙØ¹ Ù…Ù„Ù Excel Ø¨ØµÙŠØºØ© XLSX Ø£Ùˆ CSV"),400
    f=request.files["file"]
    name=(f.filename or "").lower()
    c=db()
    try:
        rows=[]
        if name.endswith(".csv"):
            import csv, io
            raw=f.read()
            text=raw.decode("utf-8-sig")
            rows=list(csv.DictReader(io.StringIO(text)))
        elif name.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                return jsonify(error="Ø§Ù„Ù…ÙƒØªØ¨Ø© openpyxl ØºÙŠØ± Ù…Ø«Ø¨ØªØ© Ø¹Ù„Ù‰ Ø§Ù„Ø³ÙŠØ±ÙØ±"),500
            wb=load_workbook(f,read_only=True,data_only=True)
            ws=wb.active
            values=list(ws.iter_rows(values_only=True))
            wb.close()
            if not values: return jsonify(error="Ù…Ù„Ù Excel ÙØ§Ø±Øº"),400
            headers=[str(x or "").strip() for x in values[0]]
            for vals in values[1:]:
                rows.append(dict(zip(headers,vals)))
        else:
            return jsonify(error="Ø§Ù„Ø§Ù…ØªØ¯Ø§Ø¯Ø§Øª Ø§Ù„Ù…Ø³Ù…ÙˆØ­Ø©: XLSX Ø£Ùˆ CSV"),400

        def norm(v):
            import unicodedata
            s=unicodedata.normalize("NFKC", str(v or "")).strip().lower()
            s=s.replace("Ø£","Ø§").replace("Ø¥","Ø§").replace("Ø¢","Ø§")
            s=s.replace("Ù‰","ÙŠ").replace("Ø©","Ù‡")
            s=re.sub(r"[\\s_\\-â€“â€”:]+", "", s)
            return s

        aliases={
            "day":["day","Ø§Ù„ÙŠÙˆÙ…","ÙŠÙˆÙ…"],
            "title":["title","Ø§Ù„Ù…Ø§Ø¯Ø©","Ø§Ù„Ù…Ø§Ø¯Ù‡","Ø§Ù„Ù…Ù‡Ù…Ø©","Ø§Ù„Ù…Ù‡Ù…Ù‡","Ø§Ù„Ù†Ø´Ø§Ø·","Ø§Ù„Ù†Ø´Ø§Ø· Ø§Ù„Ø¯Ø±Ø§Ø³ÙŠ","Ø§Ù„Ø¯Ø±Ø³","Ø¯Ø±Ø³"],
            "start":["start","Ø§Ù„Ø¨Ø¯Ø§ÙŠØ©","Ø§Ù„Ø¨Ø¯Ø§ÙŠÙ‡","ÙˆÙ‚Øª Ø§Ù„Ø¨Ø¯Ø§ÙŠØ©","ÙˆÙ‚Øª Ø§Ù„Ø¨Ø¯Ø§ÙŠÙ‡","Ø¨Ø¯Ø§ÙŠØ©","Ø¨Ø¯Ø§ÙŠÙ‡","Ù…Ù†"],
            "end":["end","Ø§Ù„Ù†Ù‡Ø§ÙŠØ©","Ø§Ù„Ù†Ù‡Ø§ÙŠÙ‡","ÙˆÙ‚Øª Ø§Ù„Ù†Ù‡Ø§ÙŠØ©","ÙˆÙ‚Øª Ø§Ù„Ù†Ù‡Ø§ÙŠÙ‡","Ù†Ù‡Ø§ÙŠØ©","Ù†Ù‡Ø§ÙŠÙ‡","Ø¥Ù„Ù‰","Ø§Ù„ÙŠ"],
            "type":["type","Ø§Ù„Ù†ÙˆØ¹","Ù†ÙˆØ¹","Ø§Ù„ØªØµÙ†ÙŠÙ"]
        }
        alias_map={norm(a):k for k,vals in aliases.items() for a in vals}
        normalized_rows=[]
        for r in rows:
            nr={}
            for k,v in r.items():
                nr[alias_map.get(norm(k),norm(k))]=v
            normalized_rows.append(nr)

        def fmt(v):
            if v is None: return ""
            try:
                from datetime import time, datetime
                if isinstance(v,time):
                    h=v.hour; m=v.minute
                    suffix="ØµØ¨Ø§Ø­Ù‹Ø§" if h<12 else "Ù…Ø³Ø§Ø¡Ù‹"
                    h12=h%12 or 12
                    return f"{h12}:{m:02d} {suffix}"
                if isinstance(v,datetime):
                    h=v.hour; m=v.minute
                    suffix="ØµØ¨Ø§Ø­Ù‹Ø§" if h<12 else "Ù…Ø³Ø§Ø¡Ù‹"
                    h12=h%12 or 12
                    return f"{h12}:{m:02d} {suffix}"
            except Exception: pass
            return str(v).strip()

        count=0
        skipped=0
        for r in normalized_rows:
            day=fmt(r.get("day"))
            title=fmt(r.get("title"))
            start_time=fmt(r.get("start"))
            end_time=fmt(r.get("end"))
            typ=fmt(r.get("type")) or "Ø¯Ø±Ø³"
            if not title:
                skipped+=1; continue
            c.execute("INSERT INTO schedules(user_id,day,title,start_time,end_time,type) VALUES(?,?,?,?,?,?)",
                      (session["uid"],day,title,start_time,end_time,typ)); count+=1
        c.commit(); return jsonify(imported=count,skipped=skipped,message=f"ØªÙ… Ø§Ø³ØªÙŠØ±Ø§Ø¯ {count} Ù…ÙˆØ¹Ø¯ Ø¨Ù†Ø¬Ø§Ø­")
    except Exception as e:
        c.rollback(); return jsonify(error=f"ØªØ¹Ø°Ø± Ù‚Ø±Ø§Ø¡Ø© Ù…Ù„Ù Excel: {e}"),400
    finally:
        c.close()

@app.get("/api/plan-template")
def plan_template():
    return send_from_directory(BASE,"Student_Weekly_Plan_Arabic.xlsx",as_attachment=True)

def ensure_admin():
    email=os.environ.get("ADMIN_EMAIL"); pw=os.environ.get("ADMIN_PASSWORD")
    if not email or not pw: return
    c=db(); u=c.execute("SELECT id FROM users WHERE email=?",(email.lower(),)).fetchone()
    if not u:
        c.execute("INSERT INTO users(name,email,password_hash,role,status,created_at) VALUES(?,?,?,?,?,?)",("Site Admin",email.lower(),generate_password_hash(pw),"admin","approved",datetime.utcnow().isoformat())); c.commit()
    c.close()
init_db(); ensure_admin()
if __name__=="__main__": app.run(host="0.0.0.0",port=int(os.environ.get("PORT","3000")))


